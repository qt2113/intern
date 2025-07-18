import os
import re
import openpyxl
from openpyxl import load_workbook
from xlrd import open_workbook
from xlutils.copy import copy

def replace_in_excel(file_path, replacements):
    _, file_extension = os.path.splitext(file_path)
    if file_extension.lower() == '.xlsx':
        wb = load_workbook(file_path)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_str = str(cell.value)
                        new_str = cell_str
                        for old_value, new_value in replacements.items():
                            if old_value in new_str:
                                new_str = new_str.replace(old_value, new_value)
                        if new_str != cell_str:
                            # 尝试转成数字
                            try:
                                new_val = float(new_str)
                                if new_val.is_integer():
                                    new_val = int(new_val)
                            except ValueError:
                                new_val = new_str
                            cell.value = new_val
        wb.save(file_path)


    elif file_extension.lower() == '.xls':
        rb = open_workbook(file_path, formatting_info=True)
        wb = copy(rb)
        for sheet_index in range(rb.nsheets):
            rb_sheet = rb.sheet_by_index(sheet_index)  #  获取原始 sheet
            ws = wb.get_sheet(sheet_index)             #  获取可写 sheet
            for row_index in range(rb_sheet.nrows):    #  使用 rb_sheet 获取 nrows
                for col_index in range(rb_sheet.ncols):
                    cell_value = rb_sheet.cell(row_index, col_index).value
                    cell_str = str(cell_value)
                    new_str = cell_str
                    for old_value, new_value in replacements.items():
                        if old_value in cell_str:
                            new_str = new_str.replace(old_value, new_value)
                    if new_str != cell_str:
                        # 如果替换后仍是数字，就转成数字
                        try:
                            new_val = float(new_str)
                            # 如果是整数就保留整数
                            if new_val.is_integer():
                                new_val = int(new_val)
                        except ValueError:
                            new_val = new_str
                        ws.write(row_index, col_index, new_val) 
            wb.save(file_path)


def replace_in_filename(file_path, replacements):
    filename = os.path.basename(file_path)
    new_filename = filename
    for old_value, new_value in replacements.items():
        new_filename = new_filename.replace(old_value, new_value)
    if new_filename != filename:
        new_file_path = os.path.join(os.path.dirname(file_path), new_filename)
        os.rename(file_path, new_file_path)
        return new_file_path
    return file_path

def main():
    folder_path = input("请输入包含Excel文件的文件夹路径: ").strip()

    # 构造替换字典
    replacements = {}
    while True:
        old = input("请输入要替换的原始值（输入'q'结束）: ").strip()
        if old.lower() == 'q':
            break
        new = input("请输入新的值: ").strip()
        replacements[old] = new

    # 扫描所有 Excel 文件
    for root, _, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            _, file_extension = os.path.splitext(file_path)
            if file_extension.lower() in ['.xls', '.xlsx']:
                new_file_path = replace_in_filename(file_path, replacements)
                print(f"处理文件: {new_file_path}")
                replace_in_excel(new_file_path, replacements)
                print(f"文件名已更新为: {os.path.basename(new_file_path)}")

    print("所有文件处理完成！")


if __name__ == "__main__":
    main()