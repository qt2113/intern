import os
import openpyxl
from openpyxl.utils import get_column_letter, datetime
import xlrd  # 新增：引入 xlrd 库以支持 .xls 文件
from xlutils.copy import copy  # 新增：导入 copy 函数以支持 .xls 文件的写操作
import xlwt


def replace_in_excel(file_path, old_value, new_value):
    _, file_extension = os.path.splitext(file_path)

    # 根据文件类型加载工作簿
    if file_extension.lower() == '.xlsx':
        wb = openpyxl.load_workbook(file_path)
        # 遍历工作簿中的每个工作表
        for sheet in wb.worksheets:
            # 使用 iter_rows() 遍历工作表中的每行
            for row in sheet.iter_rows():
                # 遍历行中的每个单元格
                for cell in row:
                    # 如果单元格的值包含需要替换的原始值，则将这部分字段替换为新值，且保留原始值的格式
                    if old_value in str(cell.value):  # 修改: 判断单元格值是否包含原始值
                        # 保留原始值的格式，仅替换字符串部分
                        if isinstance(cell.value, str):
                            cell.value = cell.value.replace(old_value, new_value)
                        elif isinstance(cell.value, (int, float)):
                            # 数字类型：先转为字符串替换，再转回数字
                            try:
                                cell.value = float(str(cell.value).replace(old_value, new_value))
                            except ValueError:
                                cell.value = str(cell.value).replace(old_value, new_value)
                        elif isinstance(cell.value, datetime.datetime):
                            # 日期类型：先转为字符串替换，再转回日期
                            try:
                                cell.value = datetime.datetime.strptime(
                                    str(cell.value).replace(old_value, new_value),
                                    "%Y-%m-%d %H:%M:%S"
                                )
                            except ValueError:
                                cell.value = str(cell.value).replace(old_value, new_value)
                        else:
                            # 其他类型：直接替换字符串部分
                            cell.value = str(cell.value).replace(old_value, new_value)
        # 保存对Excel文件的更改
        wb.save(file_path)
    elif file_extension.lower() == '.xls':
        # 使用 xlrd 加载 .xls 文件
        rb = xlrd.open_workbook(file_path, formatting_info=True)
        wb = copy(rb)  # 使用 xlutils.copy 创建可写的工作簿副本
        # 遍历工作簿中的每个工作表
        for sheet_index in range(rb.nsheets):
            ws = wb.get_sheet(sheet_index)
            rb_sheet = rb.sheet_by_index(sheet_index)  # 获取只读的工作表
            # 遍历工作表中的每行
            for row_index in range(rb_sheet.nrows):
                # 遍历行中的每个单元格
                for col_index in range(rb_sheet.ncols):
                    cell_value = rb_sheet.cell(row_index, col_index).value
                    # 如果单元格的值包含需要替换的原始值，则将这部分字段替换为新值，且保留原始值的格式
                    if old_value in str(cell_value):  # 修改: 判断单元格值是否包含原始值
                        # 保留原始值的格式，仅替换字符串部分
                        if isinstance(cell_value, str):
                            ws.write(row_index, col_index, cell_value.replace(old_value, new_value))
                        elif isinstance(cell_value, (int, float)):
                            # 数字类型：先转为字符串替换，再转回数字
                            try:
                                ws.write(row_index, col_index, float(str(cell_value).replace(old_value, new_value)))
                            except ValueError:
                                ws.write(row_index, col_index, str(cell_value).replace(old_value, new_value))
                        elif isinstance(cell_value, xlrd.xldate.XLDate):
                            # 日期类型：先转为字符串替换，再转回日期
                            try:
                                date_value = xlrd.xldate.xldate_as_datetime(cell_value, rb.datemode)
                                new_date_str = str(date_value).replace(old_value, new_value)
                                new_date_value = datetime.datetime.strptime(new_date_str, "%Y-%m-%d %H:%M:%S")
                                ws.write(row_index, col_index, new_date_value)
                            except ValueError:
                                ws.write(row_index, col_index, str(cell_value).replace(old_value, new_value))
                        else:
                            # 其他类型：直接替换字符串部分
                            ws.write(row_index, col_index, str(cell_value).replace(old_value, new_value))
        # 保存对Excel文件的更改
        wb.save(file_path)
    else:
        raise ValueError(f"不支持的文件类型: {file_extension}")


def replace_in_filename(file_path, old_value, new_value):
    """
    替换Excel文件名中的指定字段

    本函数旨在更新文件名中包含的特定文本，通过将旧文本替换为新文本，以适应文件命名的变更需求
    它首先尝试在文件路径中找到旧文本，并将其替换为新文本，然后根据新的文件名对文件进行重命名

    参数:
        file_path (str): 文件的原始路径，即需要进行名称更改的文件的路径
        old_value (str): 需要被替换的文件名中的旧文本
        new_value (str): 用于替换旧文本的新文本

    返回:
        str: 返回更新后的文件路径如果文件名未发生改变，则返回原始文件路径
    """
    # 生成新的文件路径，通过替换旧值为新值
    new_file_path = file_path.replace(old_value, new_value)

    # 检查文件路径是否发生变化，如果发生变化，则重命名文件
    if new_file_path != file_path:
        try:
            os.rename(file_path, new_file_path)
            return new_file_path
        except FileExistsError:
            print(f"文件 {new_file_path} 已经存在，无法重命名 {file_path}")
        except PermissionError:
            print(f"无法重命名 {file_path}，文件可能被其他程序锁定")
        except Exception as e:
            print(f"重命名文件时发生错误: {e}")
        return file_path

    # 如果文件路径未发生变化，即不需要重命名，直接返回原始文件路径
    return file_path


def batch_replace(directory, old_value, new_value):
    """
    批量替换指定目录下所有Excel文件中的字段和文件名

    遍历指定目录及其子目录下的所有文件，对于每个Excel文件（根据文件扩展名判断），
    首先替换其文件名中的指定字段，然后替换文件内容中的指定字段

    参数:
    directory: str, 要处理的目录路径
    old_value: str, 需要被替换的旧字段
    new_value: str, 替换旧字段的新字段
    """
    # 使用os.walk遍历目录及其子目录
    for root, dirs, files in os.walk(directory):
        for file in files:
            # 检查文件是否为Excel文件
            if file.endswith(('.xlsx', '.xls')):
                file_path = os.path.join(root, file)
                # 替换文件名中的字段
                new_file_path = replace_in_filename(file_path, old_value, new_value)
                # 替换Excel文件中的字段
                try:
                    replace_in_excel(new_file_path, old_value, new_value)
                except PermissionError:
                    print(f"无法修改 {new_file_path}，文件可能被其他程序锁定")
                except Exception as e:
                    print(f"修改文件 {new_file_path} 时发生错误: {e}")


if __name__ == "__main__":
    directory = input("请输入Excel文件所在的目录路径: ")
    old_value = input("请输入需要替换的字段: ")
    new_value = input("请输入替换为的新字段: ")

    batch_replace(directory, old_value, new_value)
    print(batch_replace(directory, old_value, new_value))
    print("替换完成！")
